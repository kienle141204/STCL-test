"""
generate_results_excel.py
─────────────────────────
Đọc file CSV kết quả thực nghiệm, tính trung bình theo
(method × backbone × dataset), rồi xuất ra file Excel
với đầy đủ pivot tables và tô màu best/2nd-best.

Cách dùng:
    python generate_results_excel.py --input res.csv --output results.xlsx

Yêu cầu:
    pip install pandas openpyxl
"""

import argparse
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Cấu hình thứ tự hiển thị ──────────────────────────────────────────────
METHOD_ORDER   = {'EAC': 0, 'RAP': 1, 'KPrompt': 2}
DATASET_ORDER  = {'AIR': 0, 'ENERGY-Wind': 1, 'PEMS': 2}
BACKBONE_ORDER = {'stgnn': 0, 'dcrnn': 1, 'astgnn': 2, 'tgcn': 3}

METHODS   = ['EAC', 'RAP', 'KPrompt']
DATASETS  = ['AIR', 'ENERGY-Wind', 'PEMS']
BACKBONES = ['stgnn', 'dcrnn', 'astgnn', 'tgcn']
METRICS   = ['avg_MAE', 'avg_RMSE', 'avg_MAPE']

# ── Màu sắc ───────────────────────────────────────────────────────────────
METHOD_COLORS = {
    'EAC':     {'light': 'D6EAF8', 'dark': '2471A3', 'mid': 'AED6F1'},
    'RAP':     {'light': 'D5F5E3', 'dark': '1E8449', 'mid': 'A9DFBF'},
    'KPrompt': {'light': 'FDEBD0', 'dark': 'CA6F1E', 'mid': 'F0B27A'},
}
DATASET_DARK  = {'AIR': '2471A3', 'ENERGY-Wind': '1E8449', 'PEMS': 'CA6F1E'}

BEST_FILL   = 'C0392B'   # đỏ  – tốt nhất
SECOND_FILL = '27AE60'   # xanh – tốt thứ 2
WHITE_FONT  = Font(name='Arial', size=10, bold=True, color='FFFFFF')
NORMAL_FONT = Font(name='Arial', size=10)

thin = Side(style='thin', color='BBBBBB')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


# ══════════════════════════════════════════════════════════════════════════
# Utilities
# ══════════════════════════════════════════════════════════════════════════

def _cell(ws, row, col, value=None, bg='FFFFFF', font=None,
          fmt=None, h_align='center', merge_end_col=None):
    """Ghi và định dạng 1 ô (hoặc merge nhiều ô cùng hàng)."""
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=merge_end_col)
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = PatternFill('solid', start_color=bg)
    c.font      = font or NORMAL_FONT
    c.alignment = Alignment(horizontal=h_align, vertical='center')
    c.border    = BORDER
    if fmt:
        c.number_format = fmt
    return c


def apply_rank_colors(cells_by_method):
    """
    cells_by_method: dict {method_name: openpyxl_cell}
    Tô đỏ ô có giá trị nhỏ nhất (tốt nhất),
    tô xanh ô có giá trị nhỏ thứ 2.
    """
    valid = {m: c for m, c in cells_by_method.items() if c.value is not None}
    if len(valid) < 2:
        return
    ranked = sorted(valid, key=lambda m: valid[m].value)
    for rank, m in enumerate(ranked):
        if rank == 0:
            valid[m].fill = PatternFill('solid', start_color=BEST_FILL)
            valid[m].font = WHITE_FONT
        elif rank == 1:
            valid[m].fill = PatternFill('solid', start_color=SECOND_FILL)
            valid[m].font = WHITE_FONT


def hdr(ws, row, col, value, bg='2C3E50', size=10, merge_end_col=None):
    return _cell(ws, row, col, value, bg=bg,
                 font=Font(name='Arial', bold=True, color='FFFFFF', size=size),
                 merge_end_col=merge_end_col)


# ══════════════════════════════════════════════════════════════════════════
# Sheet builders
# ══════════════════════════════════════════════════════════════════════════

def build_raw_sheet(wb, df):
    ws = wb.active
    ws.title = 'Raw Data'

    raw_cols   = ['method', 'backbone', 'dataset', 'seed',
                  'avg_MAE', 'avg_RMSE', 'avg_MAPE', 'total_time_s']
    col_labels = ['Method', 'Backbone', 'Dataset', 'Seed',
                  'avg_MAE', 'avg_RMSE', 'avg_MAPE', 'avg_time_s']

    df_raw = df[raw_cols].copy()
    df_raw = df_raw.sort_values(
        ['method', 'dataset', 'backbone', 'seed'],
        key=lambda x:
            x.map(METHOD_ORDER)   if x.name == 'method'   else
            x.map(DATASET_ORDER)  if x.name == 'dataset'  else
            x.map(BACKBONE_ORDER) if x.name == 'backbone' else x
    ).reset_index(drop=True)

    for c, lbl in enumerate(col_labels, 1):
        hdr(ws, 1, c, lbl, bg='1A252F', size=10)

    for r_idx, row in df_raw.iterrows():
        bg = METHOD_COLORS[row['method']]['light']
        for c_idx, col in enumerate(raw_cols, 1):
            val = row[col]
            fmt = '0.0000' if isinstance(val, float) else None
            _cell(ws, r_idx + 2, c_idx,
                  round(val, 4) if isinstance(val, float) else val,
                  bg=bg, fmt=fmt)

    for w, col in zip([10, 10, 14, 6, 12, 12, 12, 12], range(1, 9)):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 20


def build_summary_sheet(wb, grp):
    ws = wb.create_sheet('Summary Averages')

    ws.merge_cells('A1:H1')
    t = ws['A1']
    t.value     = 'Average Results  ·  Method × Backbone × Dataset  (3 seeds each)'
    t.font      = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    t.fill      = PatternFill('solid', start_color='1A252F')
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    for c, h in enumerate(
        ['Method', 'Backbone', 'Dataset',
         'avg_MAE', 'avg_RMSE', 'avg_MAPE (%)', 'avg_time_s', '#Seeds'], 1
    ):
        hdr(ws, 2, c, h)
    ws.row_dimensions[2].height = 20

    for r_idx, row in grp.iterrows():
        m  = row['method']
        bg = METHOD_COLORS[m]['light']
        vals = [m, row['backbone'], row['dataset'],
                round(row['avg_MAE'], 4), round(row['avg_RMSE'], 4),
                round(row['avg_MAPE'], 4), round(row['avg_time_s'], 4),
                int(row['n'])]
        r = r_idx + 3
        for c_idx, v in enumerate(vals, 1):
            fmt  = '0.0000' if isinstance(v, float) else None
            bold = c_idx <= 3
            c = ws.cell(row=r, column=c_idx, value=v)
            c.font      = Font(name='Arial', size=10, bold=bold)
            c.fill      = PatternFill('solid', start_color=bg)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = BORDER
            if fmt:
                c.number_format = fmt

    for w, col in zip([10, 10, 14, 13, 13, 13, 13, 8], range(1, 9)):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = 'A3'


def build_metric_pivot(wb, grp, metric, sheet_name):
    """
    Pivot: hàng = Dataset, cột = Backbone × Method.
    Tô màu: trong mỗi (dataset, backbone) so sánh 3 method.
    """
    ws = wb.create_sheet(sheet_name)
    n_cols = 1 + len(BACKBONES) * len(METHODS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    hdr(ws, 1, 1,
        f'{metric}  —  Dataset × Backbone × Method',
        bg='1A252F', size=12)
    ws.row_dimensions[1].height = 26

    # Hàng 2: Backbone (span 3 method)
    hdr(ws, 2, 1, '', bg='2C3E50')
    col = 2
    for bb in BACKBONES:
        hdr(ws, 2, col, bb, bg='34495E',
            merge_end_col=col + len(METHODS) - 1)
        col += len(METHODS)

    # Hàng 3: Method
    hdr(ws, 3, 1, '', bg='2C3E50')
    col = 2
    for bb in BACKBONES:
        for m in METHODS:
            hdr(ws, 3, col, m, bg=METHOD_COLORS[m]['dark'], size=9)
            col += 1

    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 18

    # Dữ liệu + tô màu
    for ds_idx, ds in enumerate(DATASETS):
        r = 4 + ds_idx
        hdr(ws, r, 1, ds, bg=DATASET_DARK[ds])
        col = 2
        for bb in BACKBONES:
            method_cells = {}
            for m in METHODS:
                sub = grp[(grp['method'] == m) &
                          (grp['backbone'] == bb) &
                          (grp['dataset']  == ds)]
                val = round(sub[metric].values[0], 4) if len(sub) else None
                c = ws.cell(row=r, column=col)
                c.value     = val
                c.font      = NORMAL_FONT
                c.fill      = PatternFill('solid',
                                start_color=METHOD_COLORS[m]['light'])
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border    = BORDER
                if val is not None:
                    c.number_format = '0.0000'
                method_cells[m] = c
                col += 1
            apply_rank_colors(method_cells)

    ws.column_dimensions['A'].width = 14
    for i in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 11


def build_backbone_pivot(wb, grp, backbone, sheet_name):
    """
    Pivot: hàng = Dataset, cột = Method × Metric.
    Tô màu: trong mỗi (dataset, metric) so sánh 3 method.
    """
    ws = wb.create_sheet(sheet_name)
    metric_labels = ['MAE', 'RMSE', 'MAPE']
    n_cols = 1 + len(METHODS) * len(METRICS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    hdr(ws, 1, 1,
        f'Backbone: {backbone}  —  Dataset × Method × Metric',
        bg='1A252F', size=12)
    ws.row_dimensions[1].height = 26

    # Hàng 2: Method (span 3 metric)
    hdr(ws, 2, 1, '', bg='2C3E50')
    col = 2
    for m in METHODS:
        hdr(ws, 2, col, m,
            bg=METHOD_COLORS[m]['dark'],
            merge_end_col=col + len(METRICS) - 1)
        col += len(METRICS)

    # Hàng 3: Metric
    hdr(ws, 3, 1, '', bg='2C3E50')
    col = 2
    for m in METHODS:
        for lbl in metric_labels:
            hdr(ws, 3, col, lbl,
                bg=METHOD_COLORS[m]['mid'], size=9)
            col += 1

    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 18

    # Dữ liệu + tô màu
    for ds_idx, ds in enumerate(DATASETS):
        r = 4 + ds_idx
        hdr(ws, r, 1, ds, bg=DATASET_DARK[ds])
        col = 2

        metric_cells = {met: {} for met in METRICS}
        for m in METHODS:
            sub = grp[(grp['method']   == m) &
                      (grp['backbone'] == backbone) &
                      (grp['dataset']  == ds)]
            for met_idx, met in enumerate(METRICS):
                val = round(sub[met].values[0], 4) if len(sub) else None
                c = ws.cell(row=r, column=col + met_idx)
                c.value     = val
                c.font      = NORMAL_FONT
                c.fill      = PatternFill('solid',
                                start_color=METHOD_COLORS[m]['light'])
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border    = BORDER
                if val is not None:
                    c.number_format = '0.0000'
                metric_cells[met][m] = c
            col += len(METRICS)

        for met in METRICS:
            apply_rank_colors(metric_cells[met])

    ws.column_dimensions['A'].width = 14
    for i in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 11


# ══════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════

def main(input_path: str, output_path: str):
    # 1. Đọc và làm sạch dữ liệu
    df = pd.read_csv(input_path)
    df['dataset'] = df['dataset'].str.strip().str.replace(' ', '')  # fix typo "PE MS"

    # 2. Tính trung bình theo (method, backbone, dataset)
    grp = (
        df.groupby(['method', 'backbone', 'dataset'], sort=False)
        .agg(
            n         = ('seed',         'count'),
            avg_MAE   = ('avg_MAE',      'mean'),
            avg_RMSE  = ('avg_RMSE',     'mean'),
            avg_MAPE  = ('avg_MAPE',     'mean'),
            avg_time_s= ('total_time_s', 'mean'),
        )
        .reset_index()
    )

    # Sắp xếp
    grp['_m'] = grp['method'].map(METHOD_ORDER)
    grp['_d'] = grp['dataset'].map(DATASET_ORDER)
    grp['_b'] = grp['backbone'].map(BACKBONE_ORDER)
    grp = (grp.sort_values(['_m', '_d', '_b'])
               .drop(columns=['_m', '_d', '_b'])
               .reset_index(drop=True))

    # 3. Tạo workbook và từng sheet
    wb = Workbook()

    build_raw_sheet(wb, df)
    build_summary_sheet(wb, grp)

    # # Pivot theo metric (MAE / RMSE / MAPE)
    # for metric in METRICS:
    #     label = metric.replace('avg_', '')          # "MAE", "RMSE", "MAPE"
    #     build_metric_pivot(wb, grp, metric, f'Pivot – {label}')

    # Pivot theo backbone (stgnn / dcrnn / astgnn / tgcn)
    for bb in BACKBONES:
        build_backbone_pivot(wb, grp, bb, f'Pivot – {bb}')

    # 4. Lưu file
    wb.save(output_path)
    print(f"✓ Đã lưu: {output_path}")
    print(f"  • {len(df)} dòng raw data")
    print(f"  • {len(grp)} nhóm trung bình")
    print(f"  • {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Tính trung bình kết quả thực nghiệm và xuất Excel.')
    parser.add_argument('--input',  default='res_2.csv',
                        help='Đường dẫn file CSV đầu vào')
    parser.add_argument('--output', default='results.xlsx',
                        help='Đường dẫn file Excel đầu ra')
    args = parser.parse_args()
    main(args.input, args.output)